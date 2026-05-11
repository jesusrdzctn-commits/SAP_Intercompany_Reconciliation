import os
import time
import json
import shutil
import pandas as pd
import win32com.client as win32
from tkinter import messagebox
from DescargaSAP import FBL1N_Intercompañias, ZFIQ02_Intercompañias, FBL3N, FBL5_Intercompañias
from Consolidacion_V2 import ejecutar_consolidacion_por_sociedad


class IntercompaniasController:
    """Controller class that handles the business logic for the Intercompañías application"""

    # Valores por defecto para los tamaños de bloque de FBL3N.
    # Solo se usan si config no trae los valores (fallback defensivo).
    _BLOQUE_PROV_DEFAULT = 2500
    _BLOQUE_CLI_DEFAULT  = 500

    def __init__(self, gui):
        self.gui = gui
        self.gui.on_download       = self.execute_download
        self.gui.on_consolidation  = self.execute_consolidation
        self.gui.on_download_large = self.execute_download_large

    # =========================================================
    # DESCARGA NORMAL
    # =========================================================
    def execute_download(self):
        """Execute the document download process from SAP"""
        config = self.gui.get_config()

        if not config['sociedades']:
            messagebox.showerror("Error", "Debe agregar al menos una sociedad")
            return

        if not self.gui.validate_dates():
            return

        confirm = messagebox.askyesno(
            "Confirmar",
            f"¿Desea iniciar la descarga para {len(config['sociedades'])} sociedad(es)?\n"
            f"Periodo: {config['date_from']} - {config['date_to']}\n\n"
            f"Los archivos de proveedores se guardarán en:\n{config['input_path']}\n\n"
            f"Los archivos de clientes se guardarán en:\n{config['clientes_path']}"
        )
        if not confirm:
            return

        os.makedirs(config['input_path'],   exist_ok=True)
        os.makedirs(config['clientes_path'], exist_ok=True)

        self.gui.disable_buttons()
        self.gui.set_status("⏳ Procesando... Por favor espere")

        try:
            self._run_download_process(config)
            self.gui.set_status("✅ ¡Proceso completado exitosamente!")
            messagebox.showinfo(
                "Éxito",
                "La descarga de documentos se completó correctamente.\n\n"
                f"Archivos de proveedores guardados en:\n{config['input_path']}\n\n"
                f"Archivos de clientes guardados en:\n{config['clientes_path']}"
            )
        except Exception as e:
            self.gui.set_status("❌ Error en el proceso")
            messagebox.showerror("Error", f"Ocurrió un error durante el proceso:\n\n{str(e)}")
        finally:
            self.gui.enable_buttons()
            if "completado exitosamente" not in self.gui.status_var.get():
                self.gui.set_status("✓ Listo para comenzar")

    def _run_download_process(self, config):
        """Proceso de descarga normal (sin chunking de documentos)."""
        try:
            excel = win32.GetObject(Class="Excel.Application")
            excel.Interactive = False
        except Exception:
            pass

        try:
            DateFrom     = config['date_from']
            DateTo       = config['date_to']
            FolderPath   = config['input_path']
            ClientesPath = config['clientes_path']
            sociedades   = config['sociedades']
            fbl1n_from   = config['fbl1n_range_from']
            fbl1n_to     = config['fbl1n_range_to']
            fbl5n_from   = config['fbl5n_range_from']
            fbl5n_to     = config['fbl5n_range_to']

            for idx, sociedad in enumerate(sociedades, 1):
                self.gui.set_status(f"📥 Procesando sociedad {sociedad} ({idx}/{len(sociedades)})...")
                self._descargar_sociedad(
                    sociedad, DateFrom, DateTo,
                    FolderPath, ClientesPath,
                    fbl1n_from, fbl1n_to, fbl5n_from, fbl5n_to,
                    idx, len(sociedades),
                )
        finally:
            try:
                excel = win32.GetObject(Class="Excel.Application")
                excel.Interactive = True
            except Exception:
                pass

    def _descargar_sociedad(
        self, sociedad, DateFrom, DateTo,
        FolderPath, ClientesPath,
        fbl1n_from, fbl1n_to, fbl5n_from, fbl5n_to,
        idx, total,
    ):
        """
        Descarga todos los archivos SAP para una sociedad en el periodo dado.
        Retorna (fbl1n_con_datos, fbl5n_con_datos).
        """
        # ── FBL1N ────────────────────────────────────────────────────
        self.gui.set_status(f"📥 Descargando FBL1N - {sociedad}...")
        FBL1_FileName = f"FBL1_Proveedores_{sociedad}.xlsx"
        fbl1n_con_datos = FBL1N_Intercompañias(
            [sociedad], DateFrom, DateTo, FolderPath, FBL1_FileName, fbl1n_from, fbl1n_to
        )
        FBL1_File = os.path.join(FolderPath, FBL1_FileName)

        time.sleep(10)
        self._cerrar_workbook_excel(FBL1_File)

        # ── ZFIQ02 ───────────────────────────────────────────────────
        self.gui.set_status(f"📥 Descargando ZFIQ02 - {sociedad}...")
        ZFIQ02_FileName = f"ZFIQ02_Proveedores_{sociedad}.xlsx"
        ZFIQ02_File     = os.path.join(FolderPath, ZFIQ02_FileName)
        ZFIQ02_Intercompañias([sociedad], ZFIQ02_File)

        # ── FBL3N Proveedores ─────────────────────────────────────────
        FBL3N_FileName = f"FBL3N_Proveedores_{sociedad}.xlsx"
        FBL3N_File     = os.path.join(FolderPath, FBL3N_FileName)

        if fbl1n_con_datos:
            self.gui.set_status(f"📄 Procesando documentos - {sociedad}...")
            df_FBL1      = pd.read_excel(FBL1_File, engine='openpyxl')
            colNDocument = df_FBL1.columns[6]
            resultado    = df_FBL1[colNDocument].dropna().astype(str).unique().tolist()

            self.gui.set_status(f"📥 Descargando FBL3N - {sociedad}...")
            if os.path.exists(FBL3N_File):
                os.remove(FBL3N_File)
            FBL3N(resultado, [sociedad], DateFrom, DateTo, FolderPath, FBL3N_FileName)
            time.sleep(8)
            self._cerrar_workbook_excel(FBL3N_File)
        else:
            self.gui.set_status(f"⚠️ FBL1N sin movimientos - {sociedad}, se omite FBL3N Proveedores")
            self._crear_excel_vacio(FBL3N_File)

        self.gui.set_status(f"✅ Proveedores {sociedad} completados ({idx}/{total})")
        time.sleep(5)

        # ── FBL5N ────────────────────────────────────────────────────
        self.gui.set_status(f"📥 Descargando FBL5N - {sociedad}...")
        FBL5_FileName   = f"FBL5N_Clientes_{sociedad}.xlsx"
        FBL5_File       = os.path.join(ClientesPath, FBL5_FileName)
        fbl5n_con_datos = FBL5_Intercompañias(
            [sociedad], DateFrom, DateTo, ClientesPath, FBL5_FileName, fbl5n_from, fbl5n_to
        )

        time.sleep(10)
        self._cerrar_workbook_excel(FBL5_File)

        # ── FBL3N Clientes ────────────────────────────────────────────
        FBL3N_Clientes_FileName = f"FBL3N_Clientes_{sociedad}.xlsx"
        FBL3N_Clientes_File     = os.path.join(ClientesPath, FBL3N_Clientes_FileName)

        if fbl5n_con_datos:
            self.gui.set_status(f"📄 Procesando documentos clientes - {sociedad}...")
            df_FBL5          = pd.read_excel(FBL5_File, engine='openpyxl')
            colNDocument_cli = df_FBL5.columns[8]
            resultado_cli    = df_FBL5[colNDocument_cli].dropna().astype(str).unique().tolist()

            self.gui.set_status(f"📥 Descargando FBL3N Clientes - {sociedad}...")
            if os.path.exists(FBL3N_Clientes_File):
                os.remove(FBL3N_Clientes_File)
            FBL3N(resultado_cli, [sociedad], DateFrom, DateTo, ClientesPath, FBL3N_Clientes_FileName)
            time.sleep(8)
            self._cerrar_workbook_excel(FBL3N_Clientes_File)
        else:
            self.gui.set_status(f"⚠️ FBL5N sin movimientos - {sociedad}, se omite FBL3N Clientes")
            self._crear_excel_vacio(FBL3N_Clientes_File)

        # ── Guardar flags ─────────────────────────────────────────────
        flags = {
            'sin_proveedores': not fbl1n_con_datos,
            'sin_clientes':    not fbl5n_con_datos,
        }
        flags_path = os.path.join(FolderPath, f"_flags_{sociedad}.json")
        with open(flags_path, 'w') as f:
            json.dump(flags, f)

        self.gui.set_status(f"✅ Sociedad {sociedad} completada ({idx}/{total})")
        time.sleep(5)

        return fbl1n_con_datos, fbl5n_con_datos

    # =========================================================
    # DESCARGA SOCIEDADES GRANDES (FBL3N en bloques de documentos)
    # =========================================================
    def execute_download_large(self):
        """
        Descarga para sociedades grandes.
        FBL1N y FBL5N se descargan completos (periodo entero).
        Solo FBL3N se divide en bloques de documentos para evitar errores de memoria en SAP.
        Los tamaños de bloque son configurables desde la GUI.
        """
        config = self.gui.get_config()

        if not config['sociedades']:
            messagebox.showerror("Error", "Debe agregar al menos una sociedad")
            return

        if not self.gui.validate_dates():
            return

        # Leer tamaños de bloque desde config (ya validados por _parse_chunk_size en la GUI)
        bloque_prov = config.get('chunk_prov', self._BLOQUE_PROV_DEFAULT)
        bloque_cli  = config.get('chunk_cli',  self._BLOQUE_CLI_DEFAULT)

        confirm = messagebox.askyesno(
            "Confirmar Descarga Sociedades Grandes",
            f"¿Desea iniciar la descarga para {len(config['sociedades'])} sociedad(es) grande(s)?\n\n"
            f"Periodo completo: {config['date_from']} - {config['date_to']}\n\n"
            f"• FBL1N y FBL5N se descargan completos.\n"
            f"• FBL3N Proveedores se divide en bloques de {bloque_prov} documentos.\n"
            f"• FBL3N Clientes se divide en bloques de {bloque_cli} documentos.\n\n"
            f"Archivos temporales en subcarpetas dentro de Input.\n"
            f"Al finalizar se apilan en Input/Proveedores e Input/Clientes."
        )
        if not confirm:
            return

        self.gui.disable_buttons()
        self.gui.set_status("⏳ Procesando... Por favor espere")

        try:
            self._run_chunked_download(config)
            self.gui.set_status("✅ ¡Descarga de sociedades grandes completada!")
            messagebox.showinfo(
                "Éxito",
                "La descarga se completó correctamente.\n\n"
                f"Archivos consolidados en:\n"
                f"  • {config['input_path']}\n"
                f"  • {config['clientes_path']}"
            )
        except Exception as e:
            self.gui.set_status("❌ Error en descarga de sociedades grandes")
            messagebox.showerror("Error", f"Ocurrió un error durante la descarga:\n\n{str(e)}")
        finally:
            self.gui.enable_buttons()
            if "completada" not in self.gui.status_var.get():
                self.gui.set_status("✓ Listo para comenzar")

    @staticmethod
    def _dividir_en_bloques(lista, tamano_bloque):
        """
        Divide una lista en sub-listas de tamaño fijo.

        Ejemplo con tamano_bloque=1000:
          lista de 2300 → [[...1000], [...1000], [...300]]
        """
        if not lista:
            return []
        return [
            lista[i:i + tamano_bloque]
            for i in range(0, len(lista), tamano_bloque)
        ]

    def _run_chunked_download(self, config):
        """
        Flujo de descarga para sociedades grandes por sociedad.
        Los tamaños de bloque se leen desde config para que el usuario
        los pueda ajustar sin tocar el código.

          PROVEEDORES
            1. FBL1N  → descarga completa (periodo entero)
            2. ZFIQ02 → descarga completa
            3. FBL3N Proveedores → bloques de chunk_prov docs únicos del FBL1N
            4. Apilar FBL3N Proveedores → archivo final

          CLIENTES
            5. FBL5N  → descarga completa (periodo entero)
            6. FBL3N Clientes → bloques de chunk_cli docs únicos del FBL5N
            7. Apilar FBL3N Clientes → archivo final
        """
        try:
            excel = win32.GetObject(Class="Excel.Application")
            excel.Interactive = False
        except Exception:
            pass

        try:
            DateFrom     = config['date_from']
            DateTo       = config['date_to']
            FolderPath   = config['input_path']
            ClientesPath = config['clientes_path']
            fbl1n_from   = config['fbl1n_range_from']
            fbl1n_to     = config['fbl1n_range_to']
            fbl5n_from   = config['fbl5n_range_from']
            fbl5n_to     = config['fbl5n_range_to']
            total_soc    = len(config['sociedades'])

            # Tamaños de bloque configurados por el usuario (con fallback defensivo)
            bloque_prov = config.get('chunk_prov', self._BLOQUE_PROV_DEFAULT)
            bloque_cli  = config.get('chunk_cli',  self._BLOQUE_CLI_DEFAULT)

            for idx_soc, sociedad in enumerate(config['sociedades'], 1):
                self.gui.set_status(
                    f"🏢 Procesando sociedad grande {sociedad} ({idx_soc}/{total_soc})..."
                )
                print(f"\n{'='*60}")
                print(f"[LARGE] Iniciando sociedad {sociedad} ({idx_soc}/{total_soc})")
                print(f"[LARGE] Bloques: Proveedores={bloque_prov} docs | Clientes={bloque_cli} docs")
                print(f"{'='*60}")

                # Carpetas temporales independientes por sociedad
                tmp_prov = os.path.join(FolderPath,   f"_chunks_{sociedad}")
                tmp_cli  = os.path.join(ClientesPath, f"_chunks_{sociedad}")
                os.makedirs(tmp_prov, exist_ok=True)
                os.makedirs(tmp_cli,  exist_ok=True)

                # Rutas finales de salida
                fbl1n_final   = os.path.join(FolderPath,   f"FBL1_Proveedores_{sociedad}.xlsx")
                zfiq02_final  = os.path.join(FolderPath,   f"ZFIQ02_Proveedores_{sociedad}.xlsx")
                fbl3n_p_final = os.path.join(FolderPath,   f"FBL3N_Proveedores_{sociedad}.xlsx")
                fbl5n_final   = os.path.join(ClientesPath, f"FBL5N_Clientes_{sociedad}.xlsx")
                fbl3n_c_final = os.path.join(ClientesPath, f"FBL3N_Clientes_{sociedad}.xlsx")

                # ==============================================
                # BLOQUE PROVEEDORES
                # ==============================================

                # ── 1. FBL1N completo ─────────────────────────────────
                self.gui.set_status(f"📥 [{sociedad}] Descargando FBL1N completo...")
                fbl1n_fname = f"FBL1_Proveedores_{sociedad}.xlsx"
                fbl1n_tmp   = os.path.join(tmp_prov, fbl1n_fname)

                print(f"[FBL1N] Descargando periodo completo {DateFrom} - {DateTo}...")
                fbl1n_ok = FBL1N_Intercompañias(
                    [sociedad], DateFrom, DateTo,
                    tmp_prov, fbl1n_fname, fbl1n_from, fbl1n_to
                )
                time.sleep(10)
                self._cerrar_workbook_excel(fbl1n_tmp)
                print(f"[FBL1N] Resultado: {fbl1n_ok}")

                # ── 2. ZFIQ02 completo ────────────────────────────────
                self.gui.set_status(f"📥 [{sociedad}] Descargando ZFIQ02...")
                zfiq02_tmp = os.path.join(tmp_prov, f"ZFIQ02_Proveedores_{sociedad}.xlsx")

                print("[ZFIQ02] Descargando catálogo de proveedores...")
                zfiq02_ok = ZFIQ02_Intercompañias([sociedad], zfiq02_tmp)
                print(f"[ZFIQ02] Resultado: {zfiq02_ok}")

                # ── 3. FBL3N Proveedores en bloques de documentos ─────
                chunks_fbl3n_p = []

                if fbl1n_ok:
                    df_fbl1  = pd.read_excel(fbl1n_tmp, engine='openpyxl')
                    col_ndoc = df_fbl1.columns[6]
                    docs_prov = (
                        df_fbl1[col_ndoc]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .unique()
                        .tolist()
                    )
                    bloques_prov    = self._dividir_en_bloques(docs_prov, bloque_prov)
                    total_bloques_p = len(bloques_prov)

                    print(
                        f"[FBL3N-P] {len(docs_prov)} documentos únicos → "
                        f"{total_bloques_p} bloques de máx {bloque_prov}"
                    )

                    for idx_bloque, bloque in enumerate(bloques_prov, 1):
                        self.gui.set_status(
                            f"📄 [{sociedad}] FBL3N Proveedores "
                            f"bloque {idx_bloque}/{total_bloques_p} "
                            f"({len(bloque)} docs)..."
                        )
                        fname_p = (
                            f"FBL3N_Proveedores_{sociedad}_"
                            f"bloque{idx_bloque:03d}.xlsx"
                        )
                        path_p = os.path.join(tmp_prov, fname_p)

                        if os.path.exists(path_p):
                            os.remove(path_p)

                        print(
                            f"[FBL3N-P] Bloque {idx_bloque}/{total_bloques_p} "
                            f"— {len(bloque)} documentos..."
                        )
                        fbl3n_p_ok = FBL3N(
                            bloque, [sociedad], DateFrom, DateTo,
                            tmp_prov, fname_p
                        )
                        time.sleep(8)
                        self._cerrar_workbook_excel(path_p)

                        if fbl3n_p_ok and os.path.exists(path_p):
                            chunks_fbl3n_p.append(path_p)
                            print(f"[FBL3N-P] Bloque {idx_bloque}: con datos → {path_p}")
                        else:
                            print(f"[FBL3N-P] Bloque {idx_bloque}: sin datos → se omite")
                else:
                    print("[FBL1N] Sin datos → se omiten FBL3N Proveedores")

                # ── 4. Apilar FBL3N Proveedores ───────────────────────
                self.gui.set_status(f"📋 [{sociedad}] Apilando FBL3N Proveedores...")
                print(f"[APILAR-P] {len(chunks_fbl3n_p)} bloques → {fbl3n_p_final}")
                self._apilar_chunks(chunks_fbl3n_p, fbl3n_p_final, sociedad, "FBL3N Proveedores")

                # Copiar FBL1N y ZFIQ02 a su destino final
                if fbl1n_ok and os.path.exists(fbl1n_tmp):
                    shutil.copy2(fbl1n_tmp, fbl1n_final)
                    print(f"[FBL1N] Copiado a destino final: {fbl1n_final}")
                else:
                    self._crear_excel_vacio(fbl1n_final)

                if zfiq02_ok and os.path.exists(zfiq02_tmp):
                    shutil.copy2(zfiq02_tmp, zfiq02_final)
                    print(f"[ZFIQ02] Copiado a destino final: {zfiq02_final}")
                else:
                    self._crear_excel_vacio(zfiq02_final)

                self.gui.set_status(f"✅ [{sociedad}] Proveedores completados.")
                print(f"[PROV] Bloque proveedores {sociedad} finalizado.")

                # Pausa extendida + reset de sesión SAP antes de iniciar clientes.
                # El primer chunk de FBL3N Clientes crasheaba porque la sesión
                # arrastraba memoria acumulada de todos los chunks de proveedores.
                # Navegar a /n limpia el estado interno de SAP antes de continuar.
                self.gui.set_status(f"⏸️ [{sociedad}] Esperando liberación de memoria SAP...")
                self._reset_sesion_sap()
                time.sleep(15)

                # ==============================================
                # BLOQUE CLIENTES
                # ==============================================

                # ── 5. FBL5N completo ─────────────────────────────────
                self.gui.set_status(f"📥 [{sociedad}] Descargando FBL5N completo...")
                fbl5n_fname = f"FBL5N_Clientes_{sociedad}.xlsx"
                fbl5n_tmp   = os.path.join(tmp_cli, fbl5n_fname)

                print(f"[FBL5N] Descargando periodo completo {DateFrom} - {DateTo}...")
                fbl5n_ok = FBL5_Intercompañias(
                    [sociedad], DateFrom, DateTo,
                    tmp_cli, fbl5n_fname, fbl5n_from, fbl5n_to
                )
                time.sleep(10)
                self._cerrar_workbook_excel(fbl5n_tmp)
                print(f"[FBL5N] Resultado: {fbl5n_ok}")

                # ── 6. FBL3N Clientes en bloques de documentos ────────
                chunks_fbl3n_c = []

                if fbl5n_ok:
                    df_fbl5      = pd.read_excel(fbl5n_tmp, engine='openpyxl')
                    col_ndoc_cli = df_fbl5.columns[8]
                    docs_cli = (
                        df_fbl5[col_ndoc_cli]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .unique()
                        .tolist()
                    )
                    bloques_cli     = self._dividir_en_bloques(docs_cli, bloque_cli)
                    total_bloques_c = len(bloques_cli)

                    print(
                        f"[FBL3N-C] {len(docs_cli)} documentos únicos → "
                        f"{total_bloques_c} bloques de máx {bloque_cli}"
                    )

                    for idx_bloque, bloque in enumerate(bloques_cli, 1):
                        self.gui.set_status(
                            f"📄 [{sociedad}] FBL3N Clientes "
                            f"bloque {idx_bloque}/{total_bloques_c} "
                            f"({len(bloque)} docs)..."
                        )
                        fname_c = (
                            f"FBL3N_Clientes_{sociedad}_"
                            f"bloque{idx_bloque:03d}.xlsx"
                        )
                        path_c = os.path.join(tmp_cli, fname_c)

                        if os.path.exists(path_c):
                            os.remove(path_c)

                        print(
                            f"[FBL3N-C] Bloque {idx_bloque}/{total_bloques_c} "
                            f"— {len(bloque)} documentos..."
                        )
                        fbl3n_c_ok = FBL3N(
                            bloque, [sociedad], DateFrom, DateTo,
                            tmp_cli, fname_c
                        )
                        time.sleep(8)
                        self._cerrar_workbook_excel(path_c)

                        if fbl3n_c_ok and os.path.exists(path_c):
                            chunks_fbl3n_c.append(path_c)
                            print(f"[FBL3N-C] Bloque {idx_bloque}: con datos → {path_c}")
                        else:
                            print(f"[FBL3N-C] Bloque {idx_bloque}: sin datos → se omite")
                else:
                    print("[FBL5N] Sin datos → se omite FBL3N Clientes")

                # ── 7. Apilar FBL3N Clientes ──────────────────────────
                self.gui.set_status(f"📋 [{sociedad}] Apilando FBL3N Clientes...")
                print(f"[APILAR-C] {len(chunks_fbl3n_c)} bloques → {fbl3n_c_final}")
                self._apilar_chunks(chunks_fbl3n_c, fbl3n_c_final, sociedad, "FBL3N Clientes")

                # Copiar FBL5N a su destino final
                if fbl5n_ok and os.path.exists(fbl5n_tmp):
                    shutil.copy2(fbl5n_tmp, fbl5n_final)
                    print(f"[FBL5N] Copiado a destino final: {fbl5n_final}")
                else:
                    self._crear_excel_vacio(fbl5n_final)

                # ── Guardar flags ─────────────────────────────────────
                flags = {
                    'sin_proveedores': not fbl1n_ok,
                    'sin_clientes':    not fbl5n_ok,
                }
                flags_path = os.path.join(FolderPath, f"_flags_{sociedad}.json")
                with open(flags_path, 'w') as f:
                    json.dump(flags, f)
                print(f"[FLAGS] Guardado: {flags}")

                # ── Limpiar carpetas temporales ───────────────────────
                self.gui.set_status(f"🧹 [{sociedad}] Limpiando archivos temporales...")
                shutil.rmtree(tmp_prov, ignore_errors=True)
                shutil.rmtree(tmp_cli,  ignore_errors=True)
                print("[CLEAN] Carpetas temporales eliminadas.")

                self.gui.set_status(
                    f"✅ [{sociedad}] ¡Completado! ({idx_soc}/{total_soc})"
                )
                time.sleep(3)

        finally:
            try:
                excel = win32.GetObject(Class="Excel.Application")
                excel.Interactive = True
            except Exception:
                pass

    @staticmethod
    def _apilar_chunks(rutas_chunks, ruta_final, sociedad, nombre_reporte):
        """
        Lee todos los archivos de bloque, los concatena en un solo DataFrame
        y lo guarda en ruta_final. Si no hay bloques con datos, crea un archivo vacío.
        """
        if not rutas_chunks:
            import openpyxl as _oxl
            wb = _oxl.Workbook()
            wb.active.title = "Sin datos"
            wb.save(ruta_final)
            print(f"[APILAR] {nombre_reporte} {sociedad}: sin datos → archivo vacío creado.")
            return

        dfs = []
        for ruta in rutas_chunks:
            try:
                df = pd.read_excel(ruta, dtype=str, engine='openpyxl')
                # Ignorar archivos marcados como "Sin datos"
                if df.empty or (len(df.columns) == 1 and "Sin datos" in df.columns[0]):
                    continue
                dfs.append(df)
            except Exception as e:
                print(f"[APILAR] Advertencia al leer bloque {ruta}: {e}")

        if not dfs:
            import openpyxl as _oxl
            wb = _oxl.Workbook()
            wb.active.title = "Sin datos"
            wb.save(ruta_final)
            print(f"[APILAR] {nombre_reporte} {sociedad}: todos los bloques vacíos → archivo vacío creado.")
            return

        df_total = pd.concat(dfs, ignore_index=True)
        df_total.to_excel(ruta_final, index=False, engine='openpyxl')
        print(f"[APILAR] {nombre_reporte} {sociedad}: {len(dfs)} bloques apilados → {ruta_final}")

    # =========================================================
    # CONSOLIDACIÓN
    # =========================================================
    def execute_consolidation(self):
        """Execute the consolidation process"""
        config = self.gui.get_config()

        mode       = config.get("consolidation_mode", "manual")
        mode_label = "Manual" if mode == "manual" else "Automático"

        confirm = messagebox.askyesno(
            "Confirmar Consolidación",
            f"¿Desea ejecutar el proceso de consolidación?\n\n"
            f"Modo de cuentas: {mode_label}\n\n"
            f"Se leerán archivos de proveedores de:\n{config['input_path']}\n"
            f"Se leerán archivos de clientes de:\n{config['clientes_path']}\n\n"
            f"El archivo consolidado se guardará en:\n{config['output_path']}"
        )
        if not confirm:
            return

        os.makedirs(config['output_path'], exist_ok=True)

        self.gui.disable_buttons()
        self.gui.set_status("⏳ Consolidando... Por favor espere")

        try:
            self._run_consolidation_process(config)
            self.gui.set_status("✅ ¡Consolidación completada!")
            messagebox.showinfo(
                "Éxito",
                "El proceso de consolidación se completó correctamente.\n\n"
                f"Archivo guardado en:\n{config['output_path']}"
            )
        except FileNotFoundError as e:
            self.gui.set_status("❌ Archivos no encontrados")
            messagebox.showerror(
                "Error - Archivos Faltantes",
                "No se encontraron todos los archivos necesarios.\n\n"
                "Asegúrese de ejecutar primero 'Descarga de Documentos'\n"
                "y que todos los archivos estén en la carpeta de entrada.\n\n"
                f"Detalles: {str(e)}"
            )
        except Exception as e:
            self.gui.set_status("❌ Error en consolidación")
            messagebox.showerror("Error", f"Ocurrió un error durante la consolidación:\n\n{str(e)}")
        finally:
            self.gui.enable_buttons()
            if "completada" not in self.gui.status_var.get():
                self.gui.set_status("✓ Listo para comenzar")

    def _run_consolidation_process(self, config):
        ruta_input_prov = config['input_path']
        ruta_input      = os.path.dirname(ruta_input_prov)
        ruta_output     = config['output_path']

        mode                = config.get("consolidation_mode", "manual")
        cuentas_proveedores = config['cuentas_proveedores_por_sociedad'] if mode == "manual" else None
        cuentas_clientes    = config['cuentas_clientes_por_sociedad']    if mode == "manual" else None

        for sociedad in config['sociedades']:
            flags_path = os.path.join(ruta_input_prov, f"_flags_{sociedad}.json")
            flags = {'sin_proveedores': False, 'sin_clientes': False}
            if os.path.exists(flags_path):
                try:
                    with open(flags_path) as f:
                        flags = json.load(f)
                except Exception:
                    pass

            ejecutar_consolidacion_por_sociedad(
                ruta_input,
                ruta_output,
                sociedad=sociedad,
                sin_proveedores=flags['sin_proveedores'],
                sin_clientes=flags['sin_clientes'],
                callback_status=self.gui.set_status,
                cuentas_proveedores=cuentas_proveedores,
                cuentas_clientes=cuentas_clientes,
            )

        num_archivos = len(config['sociedades'])
        self.gui.set_status(f"✅ {num_archivos} archivo(s) generado(s)")

    # =========================================================
    # HELPERS
    # =========================================================
    @staticmethod
    def _cerrar_workbook_excel(ruta_archivo):
        """Cierra un workbook de Excel si está abierto."""
        try:
            excel = win32.GetObject(Class="Excel.Application")
            for wb in list(excel.Workbooks):
                try:
                    if os.path.abspath(wb.FullName) == os.path.abspath(ruta_archivo):
                        wb.Close(SaveChanges=False)
                except Exception:
                    pass
        except Exception:
            pass

    @staticmethod
    def _crear_excel_vacio(ruta_archivo):
        """Crea un archivo Excel vacío como placeholder."""
        import openpyxl as _oxl
        wb = _oxl.Workbook()
        wb.active.title = "Sin datos"
        wb.save(ruta_archivo)

    @staticmethod
    def _reset_sesion_sap():
        """
        Navega SAP a la pantalla de inicio (/n) para liberar la memoria
        acumulada entre bloques grandes de descarga. Se llama exclusivamente
        en la transición entre el bloque de proveedores y el de clientes,
        donde la sesión arrastra el estado de todos los chunks anteriores.
        Falla silenciosamente si SAP no está disponible en ese momento.
        """
        try:
            SapGuiAuto  = win32.GetObject('SAPGUI')
            application = SapGuiAuto.GetScriptingEngine
            connection  = application.Children(0)
            session     = connection.Children(0)
            session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
            session.findById("wnd[0]").sendVKey(0)
            print("[RESET-SAP] Sesión reseteada a pantalla de inicio.")
        except Exception as e:
            print(f"[RESET-SAP] No se pudo resetear la sesión (continuando): {e}")
